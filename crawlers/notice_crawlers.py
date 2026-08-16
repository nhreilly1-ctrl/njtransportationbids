"""
notice_crawlers.py
------------------
Parser implementations for all NJ public notice sources.

Each parser:
  - Receives a source dict from notice_sources.py
  - Fetches and parses HTML
  - Returns list of raw notice dicts ready for enrichment

Output notice dict shape:
  {
    id, title, notice_excerpt, source_id, source_name, source_tier,
    source_url, official_url, county, entity_type,
    notice_type, notice_subtype,
    due_date_raw, contract_number,
    access_type, platform, paywalled,
    crawled_at, raw_html_snippet
  }
"""

import io, re, hashlib, time, logging
from datetime import date, datetime, timezone
from difflib import SequenceMatcher
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pypdf import PdfReader

log = logging.getLogger(__name__)

# ── HTTP helpers ──────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/138.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

def _get(url, timeout=20):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return r
    except requests.RequestException as e:
        log.warning(f"GET {url} failed: {e}")
        return None

def _soup(html, parser="html.parser"):
    return BeautifulSoup(html, parser)

def _now():
    return datetime.now(timezone.utc).isoformat()

def _make_id(source_id, title, url=""):
    raw = f"{source_id}:{title}:{url}"
    return "notice-" + hashlib.md5(raw.encode()).hexdigest()[:12]

def _clean(text):
    """Strip excess whitespace, normalize dashes."""
    if not text: return ""
    text = re.sub(r'\s+', ' ', text).strip()
    text = re.sub(r'\s*[–—]\s*', ' - ', text)
    return text

def _excerpt(text, chars=400):
    """Return cleaned excerpt up to chars, ending at a word boundary."""
    text = _clean(text)
    if len(text) <= chars:
        return text
    cut = text[:chars].rsplit(' ', 1)[0]
    return cut + "..."

# Transportation keyword filter — applied to municipal/county generic crawls
PROFESSIONAL_SCOPE_KW = [
    "architectural", "architecture", "engineering services",
    "engineering assistance", "engineering support", "engineering design",
    "construction management", "construction inspection", "construction monitoring", "bridge inspection",
    "design services", "surveying services", "traffic engineering",
    "transportation planning", "corridor plan", "program management",
    "cost estimating", "constructability", "geotechnical", "structural evaluation",
]

CONSTRUCTION_SCOPE_KW = [
    "roadway", "road improvement", "road resurfacing", "road repair",
    "bridge", "culvert", "drainage", "pavement", "paving", "milling",
    "overlay", "curb", "sidewalk", "intersection", "signal",
    "guardrail", "guide rail", "highway", "transportation",
    "traffic", "streetscape", "resurfacing", "reconstruction",
    "maintenance contract", "job order contract", "rail rehabilitation", "rail grinding",
    "track rehabilitation", "station rehabilitation", "platform lift",
    "structural repair", "hvac overhaul", "tank installation", "ferry retrofit",
]

SCOPE_EXCLUSIONS = [
    "advertising management", "ticket stock", "cleaning services",
    "broker dealer", "towing and recovery", "office supplies",
    "software subscription", "insurance services", "legal services",
    "purchase and delivery", "parts and supplies",
]


def _classify_transport_scope(title, body=""):
    text = f"{title} {body}".lower()
    def has_keyword(keyword):
        return re.search(rf"(?<!\w){re.escape(keyword)}(?!\w)", text) is not None

    if any(has_keyword(keyword) for keyword in SCOPE_EXCLUSIONS):
        return None
    if any(has_keyword(keyword) for keyword in PROFESSIONAL_SCOPE_KW):
        return "professional_services"
    if any(has_keyword(keyword) for keyword in CONSTRUCTION_SCOPE_KW):
        return "construction"
    return None


def _is_transport_relevant(title, body=""):
    return _classify_transport_scope(title, body) is not None


def _base_record(source, title, official_url, notice_type, due_date="", contract_number="", excerpt=""):
    return {
        "id": _make_id(source["id"], title, contract_number or official_url),
        "title": title,
        "notice_excerpt": _excerpt(excerpt or title),
        "source_id": source["id"],
        "source_name": source["name"],
        "source_tier": source["source_tier"],
        "source_url": source["url"],
        "official_url": official_url,
        "county": source.get("county", "Statewide"),
        "entity_type": source["entity_type"],
        "notice_type": notice_type,
        "notice_subtype": notice_type,
        "due_date_raw": due_date,
        "contract_number": contract_number,
        "access_type": source["access_type"],
        "platform": source["platform"],
        "paywalled": False,
        "source_status": "open",
        "crawled_at": _now(),
    }


# ── NJDOT Construction ────────────────────────────────────────────────────────

def parse_njdot_construction(source):
    """
    NJDOT current advertised projects page.
    The page renders project data in an HTML table with columns:
    Contract No | Description | Counties | Let Date | Download
    """
    records = []
    r = _get(source["url"])
    if not r: return records

    soup = _soup(r.text)
    # Find all table rows — skip header
    tables = soup.find_all("table")
    for table in tables:
        rows = table.find_all("tr")
        for row in rows:
            cells = row.find_all(["td","th"])
            if len(cells) < 2: continue

            values = [_clean(cell.get_text(" ", strip=True)) for cell in cells]
            if values[0].lower() in {"letting date", "contract no", "contract number"}:
                continue
            if len(values) == 2:
                let_date = values[0]
                project_blocks = [
                    block for block in cells[1].find_all("p", recursive=False)
                    if _clean(block.get_text(" ", strip=True))
                ] or [cells[1]]

                for block in project_blocks:
                    raw_description = _clean(block.get_text(" ", strip=True))
                    if len(raw_description) < 20 or raw_description.lower().startswith("the bid date change"):
                        continue
                    if not re.search(
                        r"\b(?:contract|project|route|maintenance|bridge|pavement|paving|"
                        r"drainage|reconstruction|resurfacing|vegetation)\b",
                        raw_description,
                        re.I,
                    ):
                        continue

                    changed_date = re.search(
                        r"date for receipt of bids is changed to\s+"
                        r"((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\.?\s+\d{1,2},?\s+\d{4})",
                        raw_description,
                        re.I,
                    )
                    project_due = changed_date.group(1) if changed_date else let_date
                    description = re.sub(
                        r"\s*[-\u2013\u2014]\s*The date for receipt of Bids is CHANGED.*$",
                        "",
                        raw_description,
                        flags=re.I,
                    ).strip()
                    contract_match = re.search(r"\bContract\s+(?:No\.?|#)\s*([A-Z0-9.-]+)", description, re.I)
                    dp_match = re.search(r"\bDP\s+No\.?\s*:?\s*([A-Z0-9.-]+)", description, re.I)
                    contract_no = contract_match.group(1) if contract_match else f"DP-{dp_match.group(1)}" if dp_match else ""
                    contract_no = contract_no.rstrip(".")
                    link = block.find("a", href=True)
                    official_url = urljoin(source["url"], link["href"]) if link else source["url"]

                    records.append({
                        "id": _make_id(source["id"], description),
                        "title": description,
                        "notice_excerpt": f"NJDOT construction contract. {description}",
                        "source_id": source["id"],
                        "source_name": source["name"],
                        "source_tier": source["source_tier"],
                        "source_url": source["url"],
                        "official_url": official_url,
                        "county": "Statewide",
                        "entity_type": source["entity_type"],
                        "notice_type": "construction",
                        "notice_subtype": "construction",
                        "due_date_raw": project_due,
                        "contract_number": contract_no,
                        "access_type": source["access_type"],
                        "platform": source["platform"],
                        "paywalled": False,
                        "source_status": "open",
                        "crawled_at": _now(),
                    })
                continue
            contract_no = values[0]
            description = values[1]
            counties = values[2] or "Statewide"
            let_date = values[3] if len(values) > 3 else ""

            # Get download link
            link = cells[-1].find("a") if cells else None
            official_url = urljoin(source["url"], link["href"]) if link and link.get("href") else source["url"]

            if not description or len(description) < 10: continue

            title = description
            if contract_no:
                title = f"Contract {contract_no} — {description}"

            records.append({
                "id":             _make_id(source["id"], title),
                "title":          title,
                "notice_excerpt": f"NJDOT construction contract. {description}. Counties: {counties}.",
                "source_id":      source["id"],
                "source_name":    source["name"],
                "source_tier":    source["source_tier"],
                "source_url":     source["url"],
                "official_url":   official_url,
                "county":         counties or "Statewide",
                "entity_type":    source["entity_type"],
                "notice_type":    "construction",
                "notice_subtype": "construction",
                "due_date_raw":   let_date,
                "contract_number":contract_no,
                "access_type":    source["access_type"],
                "platform":       source["platform"],
                "paywalled":      False,
                "crawled_at":     _now(),
            })

    # NJDOT keeps an amended contract in both its original and new letting-date
    # rows. Keep the latest row and prefer its direct notice PDF.
    def record_rank(record):
        raw_date = record.get("due_date_raw", "")
        parsed_date = date.min
        for fmt in ("%m/%d/%y", "%m/%d/%Y", "%B %d, %Y", "%b %d, %Y"):
            try:
                parsed_date = datetime.strptime(raw_date.replace(".", ""), fmt).date()
                break
            except ValueError:
                continue
        has_direct_notice = record.get("official_url") != source["url"]
        return parsed_date, has_direct_notice

    unique_records = {}
    for record in records:
        contract_key = (record.get("contract_number") or "").lower()
        title_key = re.sub(r"\W+", "", record.get("title", "").lower())
        key = contract_key or title_key
        current = unique_records.get(key)
        if current is None or record_rank(record) >= record_rank(current):
            unique_records[key] = record
    records = list(unique_records.values())

    # Also crawl planned ads page
    if source.get("planned_url"):
        records += _parse_njdot_planned(source)

    log.info(f"NJDOT construction: {len(records)} records")
    return records


def _parse_njdot_planned(source):
    """Parse planned advertisement page for forward-looking notices."""
    records = []
    r = _get(source["planned_url"])
    if not r: return records

    soup = _soup(r.text)
    for row in soup.find_all("tr")[1:]:
        cells = row.find_all("td")
        if len(cells) < 2: continue
        desc      = _clean(cells[0].get_text())
        counties  = _clean(cells[1].get_text()) if len(cells) > 1 else ""
        est_date  = _clean(cells[2].get_text()) if len(cells) > 2 else ""
        if not desc or len(desc) < 8: continue
        records.append({
            "id":             _make_id(source["id"], "planned:" + desc),
            "title":          f"[Planned] {desc}",
            "notice_excerpt": f"Planned NJDOT advertisement. Est. advertisement: {est_date}. Counties: {counties}.",
            "source_id":      source["id"],
            "source_name":    source["name"],
            "source_tier":    source["source_tier"],
            "source_url":     source["planned_url"],
            "official_url":   source["planned_url"],
            "county":         counties or "Statewide",
            "entity_type":    source["entity_type"],
            "notice_type":    "construction",
            "notice_subtype": "construction",
            "due_date_raw":   est_date,
            "contract_number":"",
            "access_type":    source["access_type"],
            "platform":       source["platform"],
            "paywalled":      False,
            "is_planned":     True,
            "crawled_at":     _now(),
        })
    return records


# ── NJDOT Professional Services ───────────────────────────────────────────────

def parse_njdot_profserv(source):
    """
    NJDOT professional services solicitations table.
    Columns: TP # | Due Date | Type | Discipline | Project Description | Status
    """
    records = []
    r = _get(source["url"])
    if not r: return records

    soup = _soup(r.text)
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        for row in rows[1:]:
            cells = row.find_all(["td","th"])
            if len(cells) < 4: continue

            values = [_clean(cell.get_text(" ", strip=True)) for cell in cells]
            if values[0].lower() == "tp number":
                continue

            tp_num = values[0]
            # Current layout: TP Number | Posting Date | Project Type |
            # Project Description | Status | Due Date.
            if len(values) >= 6:
                posting_date = values[1]
                discipline = values[2]
                description = values[3]
                status = values[4]
                due_date = values[5]
            else:
                posting_date = ""
                due_date = values[1]
                discipline = values[2]
                description = values[3]
                status = values[-1]

            # Parse discipline codes like "B-1 Level A H-1 Level B"
            codes = re.findall(r'[A-Z]-\d+', discipline + " " + description)
            code_str = " · ".join(codes) if codes else discipline

            if not description or len(description) < 8: continue

            title = f"NJDOT {tp_num} — {description}" if tp_num else description
            excerpt = (
                f"NJDOT professional services solicitation. "
                f"Prequalification required: {code_str}. "
                f"Due: {due_date}. Status: {status}."
            )

            records.append({
                "id":             _make_id(source["id"], title),
                "title":          title,
                "notice_excerpt": excerpt,
                "source_id":      source["id"],
                "source_name":    source["name"],
                "source_tier":    source["source_tier"],
                "source_url":     source["url"],
                "official_url":   source["url"],
                "county":         "Statewide",
                "entity_type":    source["entity_type"],
                "notice_type":    "professional_services",
                "notice_subtype": "professional_services",
                "due_date_raw":   due_date,
                "posting_date_raw": posting_date,
                "source_status": status,
                "contract_number":tp_num,
                "prequal_codes":  codes,
                "access_type":    source["access_type"],
                "platform":       source["platform"],
                "paywalled":      False,
                "crawled_at":     _now(),
            })

    log.info(f"NJDOT prof services: {len(records)} records")
    return records


def _anticipated_period_end(raw):
    """Return an approximate period end for NJDOT seasonal schedules."""
    match = re.search(r"\b(spring|summer|fall|winter)\s+(20)?(\d{2})\b", raw or "", re.I)
    if not match:
        return None
    season = match.group(1).lower()
    year = int((match.group(2) or "20") + match.group(3))
    month_day = {
        "spring": (6, 30),
        "summer": (9, 30),
        "fall": (12, 31),
        "winter": (3, 31),
    }[season]
    return date(year, *month_day)


def parse_njdot_profserv_upcoming(source):
    """Parse NJDOT's official anticipated professional-services workbooks."""
    records = []
    seen = set()
    page = _get(source["url"])
    if not page:
        return records

    soup = _soup(page.text)
    current_descriptions = []
    current_url = source.get("current_url")
    if current_url:
        current_page = _get(current_url)
        if current_page:
            current_soup = _soup(current_page.text)
            for row in current_soup.find_all("tr"):
                cells = row.find_all("td")
                if len(cells) >= 4:
                    current_descriptions.append(_clean(cells[3].get_text(" ", strip=True)).lower())
    workbook_urls = [
        urljoin(source["url"], link["href"])
        for link in soup.find_all("a", href=True)
        if link["href"].lower().endswith(".xlsx")
    ]

    for workbook_url in workbook_urls:
        response = _get(workbook_url, timeout=30)
        if not response:
            continue
        try:
            workbook = load_workbook(io.BytesIO(response.content), data_only=True, read_only=True)
        except Exception as exc:
            log.warning(f"Unable to parse NJDOT anticipated workbook {workbook_url}: {exc}")
            continue

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() if value is not None else "" for value in row]
                if len(values) < 7:
                    continue
                work_type, description = values[2], values[3]
                county, funding, schedule = values[4], values[5], values[6]
                period_end = _anticipated_period_end(schedule)
                if not description or not period_end or period_end < date.today():
                    continue
                if work_type.lower() in {"none", "none at this time", "tbd"}:
                    continue

                key = re.sub(r"\W+", " ", f"{work_type} {description}".lower()).strip()
                if key in seen:
                    continue
                normalized_description = re.sub(r"\W+", " ", description.lower()).strip()
                if any(
                    SequenceMatcher(None, normalized_description, current).ratio() >= 0.72
                    for current in current_descriptions
                ):
                    continue
                seen.add(key)
                title = f"NJDOT upcoming: {description}"
                records.append({
                    "id": _make_id(source["id"], title),
                    "title": title,
                    "notice_excerpt": (
                        f"Anticipated NJDOT professional-services solicitation. "
                        f"Type: {work_type}. Expected posting: {schedule}. Funding: {funding}."
                    ),
                    "source_id": source["id"],
                    "source_name": source["name"],
                    "source_tier": source["source_tier"],
                    "source_url": source["url"],
                    "official_url": workbook_url,
                    "county": county or "Statewide",
                    "entity_type": source["entity_type"],
                    "notice_type": "professional_services",
                    "notice_subtype": "professional_services",
                    "due_date_raw": schedule,
                    "anticipated_date_raw": schedule,
                    "contract_number": "",
                    "access_type": source["access_type"],
                    "platform": source["platform"],
                    "paywalled": False,
                    "is_planned": True,
                    "source_status": "upcoming",
                    "crawled_at": _now(),
                })

    log.info(f"NJDOT anticipated professional services: {len(records)} records")
    return records


def parse_njdot_design_build(source):
    """Parse active and future projects from NJDOT Alternative Project Delivery."""
    records = []
    response = _get(source["url"])
    if not response:
        return records
    soup = _soup(response.text)

    for row in soup.find_all("tr"):
        cells = row.find_all("td")
        values = [_clean(cell.get_text(" ", strip=True)) for cell in cells]
        if len(values) < 3 or not values[0].upper().startswith("DB"):
            continue
        contract_match = re.match(r"DB\D{0,3}(\d+)", values[0], re.I)
        contract_no = f"DB-{contract_match.group(1)}" if contract_match else ""
        project = values[2]
        if not project:
            continue
        # A populated best-value selection means the procurement is complete.
        if len(values) >= 6 and values[5]:
            continue
        link = cells[0].find("a", href=True)
        official_url = urljoin(source["url"], link["href"]) if link else source["url"]
        title = f"NJDOT {contract_no} - {project}"
        records.append({
            "id": _make_id(source["id"], title, contract_no),
            "title": title,
            "notice_excerpt": _excerpt(f"NJDOT design-build opportunity. {values[0]}. {values[1]}"),
            "source_id": source["id"],
            "source_name": source["name"],
            "source_tier": source["source_tier"],
            "source_url": source["url"],
            "official_url": official_url,
            "county": "Statewide",
            "entity_type": source["entity_type"],
            "notice_type": "construction",
            "notice_subtype": "construction",
            "due_date_raw": "",
            "contract_number": contract_no,
            "access_type": source["access_type"],
            "platform": source["platform"],
            "paywalled": False,
            "source_status": "open",
            "crawled_at": _now(),
        })

    page_text = _clean(soup.get_text(" ", strip=True))
    future_match = re.search(
        r"Project Name:\s*(.+?)\s+Location:\s*(.+?)\s+Project Type:\s*(.+?)\s+Schedule:\s*(.+?)(?:Future|About NJDOT|$)",
        page_text,
        re.I,
    )
    if future_match:
        project, location, project_type, schedule = future_match.groups()
        schedule = re.sub(r"\s+NJDOT.*$", "", schedule, flags=re.I).strip()
        title = f"NJDOT upcoming design-build: {project}"
        records.append({
            "id": _make_id(source["id"], title),
            "title": title,
            "notice_excerpt": f"Future NJDOT design-build project. Type: {project_type}. Location: {location}.",
            "source_id": source["id"],
            "source_name": source["name"],
            "source_tier": source["source_tier"],
            "source_url": source["url"],
            "official_url": source["url"],
            "county": "Mercer" if "Mercer" in location else "Statewide",
            "entity_type": source["entity_type"],
            "notice_type": "construction",
            "notice_subtype": "construction",
            "due_date_raw": schedule,
            "anticipated_date_raw": schedule,
            "contract_number": "",
            "access_type": source["access_type"],
            "platform": source["platform"],
            "paywalled": False,
            "is_planned": True,
            "source_status": "upcoming",
            "crawled_at": _now(),
        })

    log.info(f"NJDOT design-build: {len(records)} open/upcoming records")
    return records


# ── NJTA ──────────────────────────────────────────────────────────────────────

PANYNJ_NJ_SIGNALS = (
    "newark", "ewr", "path", "holland tunnel", "lincoln tunnel",
    "george washington bridge", "gwb", "bayonne bridge", "goethals bridge",
    "outerbridge", "port newark", "port elizabeth", "new jersey",
    "new york and new jersey", "all facilities", "port authority facilities",
)


def _json_text_fragments(value):
    if isinstance(value, dict):
        for key, child in value.items():
            if key == "text" and isinstance(child, str):
                yield child
            else:
                yield from _json_text_fragments(child)
    elif isinstance(value, list):
        for child in value:
            yield from _json_text_fragments(child)


def parse_panynj(source):
    """Parse Port Authority tables from its public Adobe AEM model."""
    model_url = source["url"].replace(".html", ".model.json")
    response = _get(model_url, timeout=40)
    if not response:
        raise RuntimeError("Port Authority solicitation data could not be fetched")
    try:
        payload = response.json()
    except ValueError as exc:
        raise RuntimeError("Port Authority returned invalid solicitation data") from exc

    notice_type = source["notice_type"]
    records = []
    seen = set()
    for fragment in _json_text_fragments(payload):
        if "<table" not in fragment.lower():
            continue
        soup = _soup(fragment)
        for table in soup.find_all("table"):
            first_row = table.find("tr")
            headers = _clean(first_row.get_text(" ", strip=True)).lower() if first_row else ""
            number_header = "contract number" if notice_type == "construction" else "proposal number"
            if number_header not in headers or "due date" not in headers or "description" not in headers:
                continue

            for row in table.find_all("tr"):
                cells = row.find_all(["td", "th"])
                if len(cells) < 3:
                    continue
                values = [_clean(cell.get_text(" ", strip=True)) for cell in cells]
                contract_number = values[0]
                if not contract_number or number_header in contract_number.lower():
                    continue

                description_el = cells[2].find("p")
                description = _clean(description_el.get_text(" ", strip=True) if description_el else values[2])
                if len(description) < 12:
                    continue
                if notice_type == "construction" and not any(signal in description.lower() for signal in PANYNJ_NJ_SIGNALS):
                    continue

                due_match = re.search(
                    r"\b\d{1,2}-[A-Za-z]{3}-\d{4}\b|"
                    r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[A-Za-z]*\s+\d{1,2},?\s+\d{4}\b|"
                    r"\b\d{1,2}/\d{1,2}/\d{2,4}\b",
                    values[1],
                    re.I,
                )
                due_date = due_match.group(0) if due_match else ""
                parsed_due = None
                for fmt in ("%d-%b-%Y", "%B %d, %Y", "%b %d, %Y", "%m/%d/%Y"):
                    try:
                        parsed_due = datetime.strptime(due_date, fmt).date()
                        break
                    except ValueError:
                        continue
                if parsed_due and parsed_due < date.today():
                    continue
                if notice_type == "professional_services" and not due_date:
                    continue

                link = cells[0].find("a", href=True)
                official_url = urljoin(source["url"], link["href"]) if link else source["url"]
                key = (contract_number.lower(), description.lower())
                if key in seen:
                    continue
                seen.add(key)

                title = f"PANYNJ {contract_number} - {description}"[:300]
                records.append(_base_record(
                    source,
                    title,
                    official_url,
                    notice_type,
                    due_date=due_date,
                    contract_number=contract_number,
                    excerpt=f"Port Authority of New York and New Jersey solicitation. {description}",
                ))

    log.info(f"PANYNJ {notice_type}: {len(records)} records")
    return records


def _extract_due_date_from_text(text):
    month_date = (
        r"(?:January|February|March|April|May|June|July|August|September|"
        r"October|November|December)\s+\d{1,2},?\s+\d{4}"
    )
    patterns = [
        rf"(?:qualification|proposal|bid)?\s*due date\s*:\s*({month_date})",
        rf"no later than\s+.{{0,200}}?({month_date})",
        rf"(?:bids?|proposals?|statements?)\s+(?:are\s+)?due\s+.{{0,200}}?({month_date})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.I | re.S)
        if match:
            return _clean(match.group(1))
    return ""


def parse_drpa(source):
    """Parse transportation construction and consulting solicitations from DRPA/PATCO."""
    response = _get(source["url"], timeout=30)
    if not response:
        raise RuntimeError("DRPA/PATCO solicitation page could not be fetched")
    soup = _soup(response.text)
    records = []
    for cell in soup.find_all("td"):
        heading = cell.find("h3")
        if not heading:
            continue
        heading_text = _clean(heading.get_text(" ", strip=True))
        context = _clean(cell.get_text(" ", strip=True))
        notice_type = _classify_transport_scope(heading_text, context)
        if not notice_type:
            continue
        link = cell.find("a", href=True)
        official_url = urljoin(source["url"], link["href"]) if link else source["url"]
        detail_response = _get(official_url, timeout=30) if link else None
        detail_text = context
        if detail_response:
            detail_text = _clean(_soup(detail_response.text).get_text(" ", strip=True))
        due_date = _extract_due_date_from_text(detail_text)
        contract_match = re.search(
            r"\b(?:Contract\s+No\.?|RFP|RFQ|IFB)\s*[-:#]?\s*([A-Z0-9-]+)",
            heading_text,
            re.I,
        )
        contract_number = contract_match.group(1) if contract_match else ""
        title = re.sub(
            r"^(?:Request For (?:Proposals|Qualifications)|Advertisement for Bid)\s*--\s*",
            "",
            heading_text,
            flags=re.I,
        )
        records.append(_base_record(
            source,
            f"DRPA/PATCO - {title}"[:300],
            official_url,
            notice_type,
            due_date=due_date,
            contract_number=contract_number,
            excerpt=detail_text,
        ))

    log.info(f"DRPA/PATCO: {len(records)} scoped records")
    return records


def parse_njtpa(source):
    """Parse active NJTPA transportation planning and engineering RFP cards."""
    response = _get(source["url"], timeout=30)
    if not response:
        raise RuntimeError("NJTPA RFP page could not be fetched")
    soup = _soup(response.text)
    records = []
    for item in soup.select("a.rfp-item"):
        title_el = item.select_one(".rfp-title")
        if not title_el:
            continue
        statuses = " ".join(node.get_text(" ", strip=True) for node in item.select(".rfp-status"))
        if "active" not in statuses.lower():
            continue
        title = _clean(title_el.get_text(" ", strip=True))
        deadline = ""
        for meta in item.select(".rfp-meta-item"):
            label = meta.select_one(".label")
            if not label or "deadline" not in label.get_text(" ", strip=True).lower():
                continue
            time_el = meta.find("time")
            raw = time_el.get("datetime", "") if time_el else meta.get_text(" ", strip=True)
            match = re.search(
                r"\b(?:January|February|March|April|May|June|July|August|September|"
                r"October|November|December)\s+\d{1,2},?\s+\d{4}\b",
                raw,
                re.I,
            )
            deadline = match.group(0) if match else raw
            break
        notice = item.select_one(".rfp-card-notice")
        excerpt = _clean(notice.get_text(" ", strip=True)) if notice else title
        records.append(_base_record(
            source,
            f"NJTPA - {title}"[:300],
            urljoin(source["url"], item.get("href", "")),
            "professional_services",
            due_date=deadline,
            excerpt=f"North Jersey Transportation Planning Authority RFP. {excerpt}",
        ))

    log.info(f"NJTPA: {len(records)} active RFPs")
    return records


def parse_njta(source):
    """
    NJ Turnpike Authority current solicitations page.
    Lists construction contracts (T-series, P-series) and
    engineering professional services (OPS numbers) together.
    """
    records = []
    r = _get(source["url"])
    if not r: return records

    soup = _soup(r.text)

    # NJTA page uses structured sections with h3/h4 headers and ul/table lists
    # Strategy: find all links with context
    for item in soup.find_all(["li", "tr", "p"]):
        text = _clean(item.get_text())
        if len(text) < 15: continue

        link = item.find("a")
        official_url = urljoin(source["url"], link["href"]) if link and link.get("href") else source["url"]

        # Extract contract number
        contract_match = re.search(
            r'\b([TP]\d{3}\.\d{3,4}|[TP]-?\d{3,4}|OPS No\. [A-Z]\d+|Order [A-Z]?\d+)\b',
            text, re.I
        )
        contract_no = contract_match.group(0) if contract_match else ""

        # Extract date
        date_match = re.search(
            r'\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s+\d{4}\b'
            r'|\b\d{1,2}/\d{1,2}/\d{2,4}\b',
            text, re.I
        )
        due_date = date_match.group(0) if date_match else ""

        lower = text.lower()
        if "engineering professional services" in lower:
            notice_type = "professional_services"
        elif "construction & maintenance" in lower or "construction and maintenance" in lower:
            notice_type = "construction"
        else:
            continue

        status_match = re.search(r"\b(Open|Closed)\b", text, re.I)
        if not status_match or not due_date:
            continue
        source_status = status_match.group(1).lower() if status_match else ""
        title = re.sub(
            r"^(?:\w+\s+\d{1,2},\s+\d{4}\s+)?(?:Open|Closed)\s+",
            "",
            text,
            flags=re.I,
        )
        title = re.sub(
            r"(?:Engineering Professional Services|Construction\s*(?:&|and)\s*Maintenance).*$",
            "",
            title,
            flags=re.I,
        ).strip()[:250]
        if len(title) < 12:
            continue
        records.append({
            "id":             _make_id(source["id"], title),
            "title":          title,
            "notice_excerpt": text[:400],
            "source_id":      source["id"],
            "source_name":    source["name"],
            "source_tier":    source["source_tier"],
            "source_url":     source["url"],
            "official_url":   official_url,
            "county":         "Statewide",
            "entity_type":    source["entity_type"],
            "notice_type":    notice_type,
            "notice_subtype": notice_type,
            "due_date_raw":   due_date,
            "source_status":  source_status,
            "contract_number":contract_no,
            "access_type":    source["access_type"],
            "platform":       source["platform"],
            "paywalled":      False,
            "crawled_at":     _now(),
        })

    log.info(f"NJTA: {len(records)} records")
    return records


def _extract_calendar_title(description, contract_no):
    quoted = re.search(r'["\u201c](.+?)["\u201d]', description)
    if quoted:
        return _clean(quoted.group(1).strip(" ."))
    cleaned = re.sub(r"^(?:Electronic\s+)?(?:Bids?|Proposals?)\s+Due:?\s*", "", description, flags=re.I)
    if contract_no:
        cleaned = re.sub(re.escape(contract_no), "", cleaned, count=1, flags=re.I)
    return _clean(cleaned).strip(" ,-.")[:250]


def _parse_njtransit_upcoming_pdf(source, pdf_url):
    records = []
    response = _get(pdf_url, timeout=30)
    if not response:
        return records
    try:
        pages = PdfReader(io.BytesIO(response.content)).pages
        text = "\n\n".join(page.extract_text() or "" for page in pages)
    except Exception as exc:
        log.warning(f"Unable to parse NJ TRANSIT upcoming PDF {pdf_url}: {exc}")
        return records

    period = ""
    category = ""
    for block in re.split(r"\n\s*\n", text):
        block = _clean(block)
        if not block:
            continue
        period_match = re.search(r"Expected Advertisement:\s*(.+?)(?:Requests|Invitations|$)", block, re.I)
        if period_match:
            period = period_match.group(1).strip()
        if "Requests for Proposals" in block:
            category = "professional_services"
        if "Invitations for Bid" in block:
            category = "construction"

        bullet_match = re.search(r"(?:^|\s)[\u2022\ufffd]\s+(.+)", block)
        if not bullet_match:
            continue
        content = bullet_match.group(1).strip()
        title = re.split(r"NJ TRANSIT is seeking", content, maxsplit=1, flags=re.I)[0].strip(" .")
        if not title:
            continue
        notice_type = _classify_transport_scope(title, content)
        if notice_type not in ("construction", "professional_services"):
            continue
        if any(exclusion in content.lower() for exclusion in SCOPE_EXCLUSIONS):
            continue

        records.append({
            "id": _make_id(source["id"], "upcoming:" + title),
            "title": f"NJ TRANSIT upcoming: {title}",
            "notice_excerpt": _excerpt(content),
            "source_id": source["id"],
            "source_name": source["name"],
            "source_tier": source["source_tier"],
            "source_url": source["url"],
            "official_url": pdf_url,
            "county": source.get("county", "Statewide"),
            "entity_type": source["entity_type"],
            "notice_type": notice_type,
            "notice_subtype": notice_type,
            "due_date_raw": period,
            "anticipated_date_raw": period,
            "contract_number": "",
            "access_type": source["access_type"],
            "platform": source["platform"],
            "paywalled": False,
            "is_planned": True,
            "source_status": "upcoming",
            "crawled_at": _now(),
        })
    return records


def parse_njtransit(source):
    """Parse future bid/proposal due rows and the upcoming-opportunities PDF."""
    records = []
    response = _get(source["url"], timeout=40)
    if not response:
        return records
    soup = _soup(response.text)

    for row in soup.find_all("tr"):
        cells = row.find_all(["td", "th"])
        if len(cells) < 4:
            continue
        values = [_clean(cell.get_text(" ", strip=True)) for cell in cells]
        try:
            due = datetime.strptime(values[0], "%m/%d/%y").date()
        except ValueError:
            continue
        description, contract_no = values[2], values[3]
        if due < date.today() or not re.search(r"\b(?:electronic )?(?:bids?|proposals?) due\b", description, re.I):
            continue
        notice_type = _classify_transport_scope(description)
        if not notice_type:
            continue
        title = _extract_calendar_title(description, contract_no)
        link = cells[2].find("a", href=True)
        official_url = urljoin(source["url"], link["href"]) if link else source["url"]
        records.append({
            "id": _make_id(source["id"], title, contract_no),
            "title": f"NJ TRANSIT {contract_no} - {title}" if contract_no else f"NJ TRANSIT - {title}",
            "notice_excerpt": _excerpt(description),
            "source_id": source["id"],
            "source_name": source["name"],
            "source_tier": source["source_tier"],
            "source_url": source["url"],
            "official_url": official_url,
            "county": source.get("county", "Statewide"),
            "entity_type": source["entity_type"],
            "notice_type": notice_type,
            "notice_subtype": notice_type,
            "due_date_raw": values[0],
            "contract_number": contract_no,
            "access_type": source["access_type"],
            "platform": source["platform"],
            "paywalled": False,
            "source_status": "open",
            "crawled_at": _now(),
        })

    upcoming_link = soup.find("a", href=re.compile(r"upcoming.*\.pdf", re.I))
    if upcoming_link:
        upcoming = _parse_njtransit_upcoming_pdf(
            source, urljoin(source["url"], upcoming_link["href"])
        )
        open_titles = [record["title"].lower() for record in records]
        for candidate in upcoming:
            candidate_title = candidate["title"].replace("NJ TRANSIT upcoming: ", "").lower()
            if any(SequenceMatcher(None, candidate_title, title).ratio() >= 0.78 for title in open_titles):
                continue
            records.append(candidate)

    log.info(f"NJ TRANSIT: {len(records)} open/upcoming records")
    return records


def parse_sjta(source):
    """Parse active SJTA Bid Express solicitations and exclude goods purchases."""
    records = []
    response = _get(source["url"], timeout=30)
    if not response:
        raise RuntimeError("SJTA Bid Express page could not be fetched")
    soup = _soup(response.text)

    for row in soup.find_all("tr"):
        cells = row.find_all("td")
        if len(cells) != 2:
            continue
        title = _clean(cells[0].get_text(" ", strip=True))
        deadline = _clean(cells[1].get_text(" ", strip=True))
        date_match = re.search(r"\b\d{1,2}/\d{1,2}/\d{4}\b", deadline)
        if not title or not date_match:
            continue
        try:
            due = datetime.strptime(date_match.group(0), "%m/%d/%Y").date()
        except ValueError:
            continue
        if due < date.today():
            continue
        notice_type = _classify_transport_scope(title)
        if not notice_type:
            continue
        link = cells[0].find("a", href=True)
        official_url = urljoin(source["url"], link["href"]) if link else source["url"]
        contract_match = re.match(r"([A-Z0-9-]+)\s+", title)
        contract_no = contract_match.group(1) if contract_match else ""
        records.append({
            "id": _make_id(source["id"], title, contract_no),
            "title": f"SJTA {title}",
            "notice_excerpt": f"South Jersey Transportation Authority solicitation. Deadline: {deadline}.",
            "source_id": source["id"],
            "source_name": source["name"],
            "source_tier": source["source_tier"],
            "source_url": source["url"],
            "official_url": official_url,
            "county": source.get("county", "Atlantic"),
            "entity_type": source["entity_type"],
            "notice_type": notice_type,
            "notice_subtype": notice_type,
            "due_date_raw": date_match.group(0),
            "contract_number": contract_no,
            "access_type": source["access_type"],
            "platform": source["platform"],
            "paywalled": False,
            "source_status": "open",
            "crawled_at": _now(),
        })

    log.info(f"SJTA: {len(records)} scoped open records")
    return records


def _extract_pdf_due_date(url):
    response = _get(url, timeout=30)
    if not response:
        return ""
    try:
        pages = PdfReader(io.BytesIO(response.content)).pages[:2]
        text = " ".join(page.extract_text() or "" for page in pages)
    except Exception as exc:
        log.warning(f"Unable to read deadline PDF {url}: {exc}")
        return ""

    month_date = (
        r"(?:January|February|March|April|May|June|July|August|September|"
        r"October|November|December)\s+\d{1,2},?\s+\d{4}"
    )
    match = re.search(rf"\buntil\b.{{0,140}}?({month_date})", text, re.I | re.S)
    if not match:
        match = re.search(rf"\b(?:bids?|proposals?)\s+due\b.{{0,100}}?({month_date})", text, re.I | re.S)
    return _clean(match.group(1)) if match else ""


# ── DRJTBC ────────────────────────────────────────────────────────────────────

def parse_drjtbc(source):
    """
    DRJTBC construction notices and professional services current procurements.
    Both pages have similar structure: project title, document links, dates.
    """
    records = []
    r = _get(source["url"])
    if not r: return records

    soup = _soup(r.text)
    is_profserv = "profserv" in source["id"] or "professional" in source["url"]

    # Find project blocks — typically div or article elements with h3/h4 titles
    for block in soup.find_all(["article","div","section"], class_=re.compile(r'project|procurement|listing|item', re.I)):
        title_el = block.find(["h2","h3","h4","strong"])
        if not title_el: continue
        title = _clean(title_el.get_text())
        if len(title) < 10 or title.lower().startswith("current notice"):
            continue

        body = _clean(block.get_text())
        link = block.find("a", href=re.compile(r'\.(pdf|doc|htm)', re.I))
        official_url = urljoin(source["url"], link["href"]) if link and link.get("href") else source["url"]

        date_match = re.search(r'\b\d{1,2}/\d{1,2}/\d{2,4}\b|\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\.?\s+\d{1,2},?\s*\d{4}\b', body, re.I)
        due_date = date_match.group(0) if date_match else ""
        if not due_date and official_url.lower().endswith(".pdf"):
            due_date = _extract_pdf_due_date(official_url)

        contract_match = re.search(r'\b(?:Contract|DB|C)-?\s*[A-Z0-9]{3,10}\b', body, re.I)
        contract_no = contract_match.group(0) if contract_match else ""

        notice_type = "professional_services" if is_profserv else "construction"

        records.append({
            "id":             _make_id(source["id"], title),
            "title":          title,
            "notice_excerpt": _excerpt(body),
            "source_id":      source["id"],
            "source_name":    source["name"],
            "source_tier":    source["source_tier"],
            "source_url":     source["url"],
            "official_url":   official_url,
            "county":         source["county"],
            "entity_type":    source["entity_type"],
            "notice_type":    notice_type,
            "notice_subtype": notice_type,
            "due_date_raw":   due_date,
            "source_status":  "open",
            "contract_number":contract_no,
            "access_type":    source["access_type"],
            "platform":       source["platform"],
            "paywalled":      False,
            "crawled_at":     _now(),
        })

    # Fallback: parse paragraphs if no blocks found
    if not records:
        for p in soup.find_all("p"):
            text = _clean(p.get_text())
            if len(text) < 30 or not _is_transport_relevant(text): continue
            link = p.find("a")
            official_url = urljoin(source["url"], link["href"]) if link and link.get("href") else source["url"]
            due_date = _extract_pdf_due_date(official_url) if official_url.lower().endswith(".pdf") else ""
            records.append({
                "id":             _make_id(source["id"], text[:80]),
                "title":          text[:150],
                "notice_excerpt": _excerpt(text),
                "source_id":      source["id"],
                "source_name":    source["name"],
                "source_tier":    source["source_tier"],
                "source_url":     source["url"],
                "official_url":   official_url,
                "county":         source["county"],
                "entity_type":    source["entity_type"],
                "notice_type":    "professional_services" if is_profserv else "construction",
                "notice_subtype": "professional_services" if is_profserv else "construction",
                "due_date_raw":   due_date,
                "source_status":  "open",
                "contract_number":"",
                "access_type":    source["access_type"],
                "platform":       source["platform"],
                "paywalled":      False,
                "crawled_at":     _now(),
            })

    log.info(f"DRJTBC {source['id']}: {len(records)} records")
    return records


# ── NJ DOS Legal Notices ──────────────────────────────────────────────────────

def parse_nj_dos_legal(source):
    """
    NJ Department of State legal notices page.
    Since Mar 2026 this is the canonical statewide legal notice repository.
    Filter aggressively for transportation/construction content.
    """
    records = []
    r = _get(source["url"])
    if not r:
        raise RuntimeError("NJ Department of State notices page could not be fetched")

    soup = _soup(r.text)

    for item in soup.find_all(["li","p","div"], class_=re.compile(r'notice|item|entry', re.I)):
        text = _clean(item.get_text())
        if len(text) < 20: continue
        if not _is_transport_relevant(text): continue

        link = item.find("a")
        official_url = urljoin(source["url"], link["href"]) if link and link.get("href") else source["url"]
        title = _clean(link.get_text()) if link else text[:150]

        date_match = re.search(
            r'\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\.?\s+\d{1,2},?\s*\d{4}\b'
            r'|\b\d{1,2}/\d{1,2}/\d{2,4}\b',
            text, re.I
        )
        due_date = date_match.group(0) if date_match else ""

        # Classify notice subtype
        scoped_type = _classify_transport_scope(text)
        if scoped_type == "professional_services":
            ntype, nsub = "public_notice", "professional_services"
        elif scoped_type == "construction":
            ntype, nsub = "public_notice", "construction"
        else:
            ntype, nsub = "public_notice", None

        records.append({
            "id":             _make_id(source["id"], title),
            "title":          title or text[:150],
            "notice_excerpt": _excerpt(text),
            "source_id":      source["id"],
            "source_name":    source["name"],
            "source_tier":    source["source_tier"],
            "source_url":     source["url"],
            "official_url":   official_url,
            "county":         "Statewide",
            "entity_type":    source["entity_type"],
            "notice_type":    ntype,
            "notice_subtype": nsub,
            "due_date_raw":   due_date,
            "contract_number":"",
            "access_type":    "Public access",
            "platform":       "NJDOS legal notices portal",
            "paywalled":      False,
            "crawled_at":     _now(),
        })

    log.info(f"NJ DOS legal notices: {len(records)} records")
    return records


# ── SoS Directory (Tier 3 seed) ───────────────────────────────────────────────

def parse_sos_directory(source):
    """
    Crawl the Secretary of State statewide legal notices directory.
    Extracts all submitted entity URLs — feeds the Tier 3 municipal crawler.
    Returns a list of {entity_name, legal_notices_url} dicts, NOT notice records.
    Call this separately from notice_runner.py to seed Tier 3.
    """
    r = _get(source["url"])
    if not r: return []

    soup = _soup(r.text)
    entities = []

    # Page lists entities with their legal notice page URLs
    for row in soup.find_all(["tr","li","div"]):
        link = row.find("a", href=re.compile(r'http', re.I))
        if not link: continue
        href = link.get("href","")
        if not href or "nj.gov/state" in href: continue  # skip self-links
        name = _clean(row.get_text())[:100]
        entities.append({
            "entity_name": name,
            "legal_notices_url": href,
            "discovered_at": _now(),
        })

    log.info(f"SoS directory: {len(entities)} entity URLs discovered")
    return entities


# ── Generic HTML List (county/municipal fallback) ─────────────────────────────

def parse_generic_html_list(source):
    """
    Fallback parser for county and municipal sites.
    Finds bid/notice links and extracts title, date, URL.
    Applies transportation keyword filter.
    """
    records = []
    r = _get(source["url"])
    if not r:
        raise RuntimeError(f"{source['name']} page could not be fetched")

    soup = _soup(r.text)
    seen = set()

    # Strategy 1: find anchor tags that look like bid postings
    bid_link_patterns = re.compile(
        r'bid|notice|rfp|rfq|solicitation|procurement|advertisement|award|contract',
        re.I
    )

    for a in soup.find_all("a", href=True):
        text = _clean(a.get_text())
        href = a.get("href","")
        if not text or len(text) < 10: continue
        if not bid_link_patterns.search(text) and not bid_link_patterns.search(href): continue
        if not _is_transport_relevant(text): continue
        if text in seen: continue
        seen.add(text)

        official_url = urljoin(source["url"], href)

        # Try to get date from surrounding context
        parent = a.find_parent(["li","tr","div","p"])
        context = _clean(parent.get_text()) if parent else text
        date_match = re.search(
            r'\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\.?\s+\d{1,2},?\s*\d{4}\b'
            r'|\b\d{1,2}/\d{1,2}/\d{2,4}\b',
            context, re.I
        )
        due_date = date_match.group(0) if date_match else ""

        contract_match = re.search(r'\bBid\s+No\.?\s*[\w-]+|\bRFP\s+[\d-]+|\bBid\s+#\s*[\w-]+', context, re.I)
        contract_no = contract_match.group(0) if contract_match else ""

        # Classify
        ntype = _classify_transport_scope(text, context)
        if not ntype:
            continue

        records.append({
            "id":             _make_id(source["id"], text),
            "title":          text[:250],
            "notice_excerpt": _excerpt(context),
            "source_id":      source["id"],
            "source_name":    source["name"],
            "source_tier":    source["source_tier"],
            "source_url":     source["url"],
            "official_url":   official_url,
            "county":         source.get("county",""),
            "entity_type":    source["entity_type"],
            "notice_type":    ntype,
            "notice_subtype": ntype,
            "due_date_raw":   due_date,
            "contract_number":contract_no,
            "access_type":    source.get("access_type","Unknown"),
            "platform":       source.get("platform","Agency website"),
            "paywalled":      False,
            "crawled_at":     _now(),
        })

    # Strategy 2: look for legal notice sections added post-PL2025-c72
    legal_section = soup.find(id=re.compile(r'legal.notice|notice.bid', re.I))
    if legal_section:
        for p in legal_section.find_all("p"):
            text = _clean(p.get_text())
            if len(text) < 30 or not _is_transport_relevant(text): continue
            if text in seen: continue
            seen.add(text)
            records.append({
                "id":             _make_id(source["id"], text[:80]),
                "title":          text[:200],
                "notice_excerpt": _excerpt(text),
                "source_id":      source["id"],
                "source_name":    source["name"],
                "source_tier":    source["source_tier"],
                "source_url":     source["url"],
                "official_url":   source["url"],
                "county":         source.get("county",""),
                "entity_type":    source["entity_type"],
                "notice_type":    "public_notice",
                "notice_subtype": None,
                "due_date_raw":   "",
                "contract_number":"",
                "access_type":    source.get("access_type","Unknown"),
                "platform":       source.get("platform","Agency website"),
                "paywalled":      False,
                "crawled_at":     _now(),
            })

    log.info(f"{source['name']}: {len(records)} records")
    return records


# ── Essex County dedicated portal ─────────────────────────────────────────────

def parse_essex_county(source):
    """Essex has a dedicated procurement portal with structured JSON-like data."""
    records = []

    # Essex legal notices page is more useful than their procurement portal
    legal_url = source.get("legal_url", source["url"])
    r = _get(legal_url)
    if not r:
        r = _get(source["url"])
    if not r: return records

    soup = _soup(r.text)
    for item in soup.find_all(["li","div","article"]):
        link = item.find("a", href=True)
        if not link: continue
        text = _clean(link.get_text())
        if len(text) < 10: continue
        if not _is_transport_relevant(text): continue

        href = link.get("href","")
        official_url = urljoin(legal_url, href)
        context = _clean(item.get_text())

        date_match = re.search(r'\b\d{1,2}/\d{1,2}/\d{2,4}\b', context)
        due_date = date_match.group(0) if date_match else ""

        ntype = _classify_transport_scope(text, context)
        if not ntype:
            continue

        records.append({
            "id":             _make_id(source["id"], text),
            "title":          text[:250],
            "notice_excerpt": _excerpt(context),
            "source_id":      source["id"],
            "source_name":    source["name"],
            "source_tier":    source["source_tier"],
            "source_url":     legal_url,
            "official_url":   official_url,
            "county":         "Essex",
            "entity_type":    source["entity_type"],
            "notice_type":    ntype,
            "notice_subtype": ntype,
            "due_date_raw":   due_date,
            "contract_number":"",
            "access_type":    source["access_type"],
            "platform":       source["platform"],
            "paywalled":      False,
            "crawled_at":     _now(),
        })

    log.info(f"Essex County: {len(records)} records")
    return records


# ── Camden County dedicated portal ────────────────────────────────────────────

def parse_camden_county(source):
    """Camden has procurements.camdencounty.com — structured listing."""
    records = []
    r = _get(source["url"])
    if not r: return records

    soup = _soup(r.text)
    for item in soup.find_all(["article","div","li"], class_=re.compile(r'procurement|bid|item', re.I)):
        title_el = item.find(["h2","h3","h4","a","strong"])
        if not title_el: continue
        title = _clean(title_el.get_text())
        if len(title) < 10 or not _is_transport_relevant(title): continue

        link = item.find("a", href=True)
        official_url = urljoin(source["url"], link["href"]) if link else source["url"]
        context = _clean(item.get_text())

        date_match = re.search(r'\b\d{1,2}/\d{1,2}/\d{2,4}\b|\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\.?\s+\d{1,2},?\s*\d{4}\b', context, re.I)
        due_date = date_match.group(0) if date_match else ""

        ntype = _classify_transport_scope(title, context)
        if not ntype:
            continue

        records.append({
            "id":             _make_id(source["id"], title),
            "title":          title,
            "notice_excerpt": _excerpt(context),
            "source_id":      source["id"],
            "source_name":    source["name"],
            "source_tier":    source["source_tier"],
            "source_url":     source["url"],
            "official_url":   official_url,
            "county":         "Camden",
            "entity_type":    source["entity_type"],
            "notice_type":    ntype,
            "notice_subtype": ntype,
            "due_date_raw":   due_date,
            "contract_number":"",
            "access_type":    source["access_type"],
            "platform":       source["platform"],
            "paywalled":      False,
            "crawled_at":     _now(),
        })

    if not records:
        records = parse_generic_html_list(source)

    log.info(f"Camden County: {len(records)} records")
    return records


# ── Monmouth County ───────────────────────────────────────────────────────────

def parse_monmouth_county(source):
    """Monmouth has a searchable portal at pol.co.monmouth.nj.us."""
    records = []

    # The Monmouth portal requires POSTing a search — use generic fallback
    # but try the open bids list first
    open_bids_url = "https://pol.co.monmouth.nj.us/County/tblBids.aspx?Status=Open"
    r = _get(open_bids_url)
    if not r:
        return parse_generic_html_list(source)

    soup = _soup(r.text)
    for row in soup.find_all("tr")[1:]:
        cells = row.find_all("td")
        if len(cells) < 3: continue

        req_id    = _clean(cells[0].get_text())
        due_date  = _clean(cells[1].get_text())
        title     = _clean(cells[2].get_text())

        if not title or not _is_transport_relevant(title): continue

        link = row.find("a", href=True)
        official_url = urljoin(open_bids_url, link["href"]) if link else source["url"]

        ntype = _classify_transport_scope(title)
        if not ntype:
            continue

        records.append({
            "id":             _make_id(source["id"], title),
            "title":          f"Monmouth County — {title}",
            "notice_excerpt": f"Monmouth County bid solicitation. Request ID: {req_id}. Due: {due_date}. {title}",
            "source_id":      source["id"],
            "source_name":    source["name"],
            "source_tier":    source["source_tier"],
            "source_url":     open_bids_url,
            "official_url":   official_url,
            "county":         "Monmouth",
            "entity_type":    source["entity_type"],
            "notice_type":    ntype,
            "notice_subtype": ntype,
            "due_date_raw":   due_date,
            "contract_number":req_id,
            "access_type":    source["access_type"],
            "platform":       source["platform"],
            "paywalled":      False,
            "crawled_at":     _now(),
        })

    log.info(f"Monmouth County: {len(records)} records")
    return records


# ── Gloucester County ─────────────────────────────────────────────────────────

def parse_gloucester_county(source):
    """Gloucester uses .aspx bid listing with bidID params."""
    records = []
    r = _get(source["url"])
    if not r: return records

    soup = _soup(r.text)
    for row in soup.find_all("tr")[1:]:
        cells = row.find_all("td")
        if len(cells) < 2: continue
        title = _clean(cells[0].get_text() if cells else "")
        if not title or not _is_transport_relevant(title): continue

        link = row.find("a", href=re.compile(r'bidID', re.I))
        official_url = urljoin(source["url"], link["href"]) if link else source["url"]
        date_cell = _clean(cells[1].get_text()) if len(cells) > 1 else ""

        ntype = _classify_transport_scope(title, _clean(row.get_text()))
        if not ntype:
            continue

        records.append({
            "id":             _make_id(source["id"], title),
            "title":          f"Gloucester County — {title}",
            "notice_excerpt": _excerpt(_clean(row.get_text())),
            "source_id":      source["id"],
            "source_name":    source["name"],
            "source_tier":    source["source_tier"],
            "source_url":     source["url"],
            "official_url":   official_url,
            "county":         "Gloucester",
            "entity_type":    source["entity_type"],
            "notice_type":    ntype,
            "notice_subtype": ntype,
            "due_date_raw":   date_cell,
            "contract_number":"",
            "access_type":    source["access_type"],
            "platform":       source["platform"],
            "paywalled":      False,
            "crawled_at":     _now(),
        })

    log.info(f"Gloucester County: {len(records)} records")
    return records


# ── Municipal Tier 3 — seeded from SoS directory ─────────────────────────────

def parse_municipal_from_sos(entity_url, entity_name, county=""):
    """
    Called by the Tier 3 runner after SoS directory provides URLs.
    Generic parser for municipal legal notice pages under PL2025-c72.
    """
    source_id = "sos-" + hashlib.md5(entity_url.encode()).hexdigest()[:8]
    synthetic_source = {
        "id":          source_id,
        "name":        entity_name,
        "source_tier": "municipal",
        "url":         entity_url,
        "access_type": "Public access",
        "platform":    "Municipal website",
        "entity_type": "Municipality",
        "county":      county,
    }
    return parse_generic_html_list(synthetic_source)


# ── Dispatcher ────────────────────────────────────────────────────────────────

PARSER_MAP = {
    "njdot_construction":   parse_njdot_construction,
    "njdot_profserv":       parse_njdot_profserv,
    "njdot_profserv_upcoming": parse_njdot_profserv_upcoming,
    "njdot_design_build":  parse_njdot_design_build,
    "panynj":               parse_panynj,
    "drpa":                 parse_drpa,
    "njtpa":                parse_njtpa,
    "njta":                 parse_njta,
    "njtransit":            parse_njtransit,
    "sjta":                 parse_sjta,
    "drjtbc":               parse_drjtbc,
    "nj_dos_legal":         parse_nj_dos_legal,
    "sos_directory":        parse_sos_directory,
    "essex_county":         parse_essex_county,
    "camden_county":        parse_camden_county,
    "monmouth_county":      parse_monmouth_county,
    "gloucester_county":    parse_gloucester_county,
    "generic_html_list":    parse_generic_html_list,
    "bidnet":               parse_generic_html_list,
    "questcdn":             parse_generic_html_list,
}

def crawl_source(source, delay=1.5):
    """Crawl a single source. Returns list of notice dicts."""
    parser_name = source.get("parser","generic_html_list")
    parser_fn   = PARSER_MAP.get(parser_name, parse_generic_html_list)
    records = parser_fn(source)
    time.sleep(delay)   # polite delay between requests
    return records
