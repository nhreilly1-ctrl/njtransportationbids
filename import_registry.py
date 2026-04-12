from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.registry_sources import RegistrySource

WORKBOOK_FIELD_MAP = {
    "Source ID": "source_id",
    "Source Level": "source_level",
    "Entity Type": "entity_type",
    "Source Name": "source_name",
    "County": "county",
    "Municipality": "municipality",
    "Region": "region",
    "Coverage / Scope": "coverage_scope",
    "Priority Tier": "priority_tier",
    "Rank / Area": "rank_area",
    "Statewide DOS Directory": "statewide_dos_directory",
    "Direct Legal Notice URL": "direct_legal_notice_url",
    "Effective Notice Entry URL": "effective_notice_entry_url",
    "Primary Procurement URL": "primary_procurement_url",
    "Verification URL / County URL": "verification_url_county_url",
    "Portal Type": "portal_type",
    "Crawl Entry": "crawl_entry",
    "Verification Status": "verification_status",
    "Source Status": "source_status",
    "Refresh Cadence": "refresh_cadence",
    "Website Ready": "website_ready",
    "Parser Hint": "parser_hint",
    "Use For": "use_for",
    "Notes": "notes",
}

REQUIRED_WORKBOOK_COLUMNS = [
    "Source ID",
    "Source Level",
    "Entity Type",
    "Source Name",
    "Region",
    "Effective Notice Entry URL",
    "Primary Procurement URL",
    "Crawl Entry",
]


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _priority_rank_from_value(value: str | None) -> int | None:
    if not value:
        return None
    digits = ''.join(ch for ch in value if ch.isdigit())
    return int(digits) if digits else None


def _import_enabled(value: str | None) -> bool:
    cleaned = (value or '').strip().lower()
    return cleaned in {'yes', 'partial', 'true', '1'}


def _build_payload(row: dict[str, Any], *, source_sheet: str | None = None, source_row_number: int | None = None) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for workbook_name, model_name in WORKBOOK_FIELD_MAP.items():
        payload[model_name] = _clean(row.get(workbook_name))

    payload['priority_rank'] = _priority_rank_from_value(payload.get('rank_area'))
    payload['import_enabled'] = _import_enabled(payload.get('website_ready'))
    payload['source_sheet'] = source_sheet
    payload['source_row_number'] = source_row_number

    for col in REQUIRED_WORKBOOK_COLUMNS:
        if not _clean(row.get(col)):
            raise ValueError(f"Missing required workbook field: {col}")

    return payload


def import_registry_rows(db: Session, rows: Iterable[dict[str, Any]], *, source_sheet: str | None = None) -> int:
    count = 0
    for idx, row in enumerate(rows, start=2):
        payload = _build_payload(row, source_sheet=source_sheet, source_row_number=idx)
        db.merge(RegistrySource(**payload))
        count += 1
    db.commit()
    return count


def import_registry_csv(db: Session, csv_path: str) -> int:
    path = Path(csv_path)
    with path.open(newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        return import_registry_rows(db, reader, source_sheet='Master Registry CSV export')


def import_registry_workbook(db: Session, workbook_path: str, sheet_name: str = 'Master Registry') -> int:
    path = Path(workbook_path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else '' for v in next(rows)]
    reader_rows = [dict(zip(headers, row)) for row in rows if any(cell not in (None, '') for cell in row)]
    return import_registry_rows(db, reader_rows, source_sheet=sheet_name)


def export_master_registry_to_csv(workbook_path: str, output_csv_path: str, sheet_name: str = 'Master Registry') -> int:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v).strip() if v is not None else '' for v in rows[0]]
    out_path = Path(output_csv_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open('w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        written = 0
        for row in rows[1:]:
            if any(cell not in (None, '') for cell in row):
                writer.writerow(list(row))
                written += 1
    return written
